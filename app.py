from flask import Flask, request, jsonify, Response
from openpyxl import load_workbook
import os

app = Flask(__name__)

# Route for HTML + JS
@app.route('/')
def index():
    html = """
    <!DOCTYPE html>
    <html>
    <head>
      <title>Excel Calculator</title>
      <script>
        async function sendData() {
          const input = document.getElementById('inputValue').value;
          const response = await fetch('/calculate', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({input})
          });
          const data = await response.json();
          document.getElementById('result').innerText = data.result;
        }
      </script>
    </head>
    <body>
      <h1>Excel Calculation</h1>
      <input type="number" id="inputValue" placeholder="Enter value">
      <button onclick="sendData()">Calculate</button>
      <p>Result: <span id="result"></span></p>
    </body>
    </html>
    """
    return Response(html, mimetype='text/html')

# Route for Excel calculation
@app.route('/calculate', methods=['POST'])
def calculate():
    input_value = float(request.json['input'])

    if not os.path.exists('template.xlsx'):
        return jsonify({'error': 'Excel file missing'}), 500

    wb = load_workbook('template.xlsx')
    ws = wb.active
    ws['A1'] = input_value
    wb.save('template.xlsx')

    wb = load_workbook('template.xlsx', data_only=True)
    ws = wb.active
    result = ws['B1'].value

    return jsonify({'result': result})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000)
